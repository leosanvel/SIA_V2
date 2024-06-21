from .puestos import puestos
from flask import render_template, request, session, jsonify
from flask_login import current_user
from sqlalchemy.orm.exc import NoResultFound
from sqlalchemy import and_, func, inspect
from datetime import datetime, time
import openpyxl

from app import db
from rh.gestion_empleados.modelos.empleado import tPuesto
from catalogos.modelos.modelos import kEstatusPuesto, kVigencia, kCentroTrabajo, kCentroCostos

@puestos.route("/innovacion-normas/puestos/cargar-puestos", methods = ["POST", "GET"])
def cargar_puestos():

    return render_template("/cargar_puestos.html", title = "Cargar puestos")

@puestos.route("/innovacion-normas/puestos/cargar-archivo-puestos", methods = ["POST"])
def cargar_archivo_puestos():
    archivo = request.files.get("archivo")

    # Obtener nombre de columnas de la tabla
    columnas = inspect(tPuesto).all_orm_descriptors.keys()
    # Eliminar relaciones
    del columnas[32:44]

    # Relación de columnas de tabla y columnas de excel
    valor_columnas =[1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 32, 41, 40]
    mapeo_nombres = dict(zip(columnas, valor_columnas))

    if(archivo):
        wb = openpyxl.load_workbook(archivo)
        ws = wb.active

        datos_lista = []
        datos_dict = {}
        nuevo_puesto = None

        #idEstatusPuesto, idVigencia, idCentroTrabajo

        for row_value in range(12, ws.max_row + 1):
            for key in mapeo_nombres:
                datos_dict.update({key: ws.cell(row = row_value, column = mapeo_nombres[key]).value})
                if key == "idEstatusPuesto":
                    valor = db.session.query(kEstatusPuesto).filter_by(EstatusPuesto = datos_dict["idEstatusPuesto"]).one()
                    datos_dict["idEstatusPuesto"] = valor.idEstatusPuesto
                
                if key == "idVigencia":
                    valor = db.session.query(kVigencia).filter_by(Vigencia = datos_dict["idVigencia"]).one()
                    datos_dict["idVigencia"] = valor.idVigencia

                if key == "idCentroTrabajo":
                    valor = db.session.query(kCentroTrabajo).filter_by(CentroTrabajo = datos_dict["idCentroTrabajo"]).one()
                    datos_dict["idCentroTrabajo"] = valor.idCentroTrabajo

                if key == "idCentroCosto":
                    datos_dict["idCentroCosto"] = None

            datos_lista.append(datos_dict)

            try:
                puesto_existente = db.session.query(tPuesto).filter_by(ConsecutivoPuesto = datos_dict["ConsecutivoPuesto"]).one()
                #puesto_existente.update(**datos_dict)
                if not puesto_existente.CodigoPuesto == datos_dict["CodigoPuesto"]:
                    puesto_existente.Activo = 0
                    datos_dict["Activo"] = 1
                    nuevo_puesto = tPuesto(**datos_dict)
                    db.session.add(nuevo_puesto)

            except NoResultFound:
                datos_dict["Activo"] = 1
                nuevo_puesto = tPuesto(**datos_dict)
                db.session.add(nuevo_puesto)

        db.session.commit()

        CentroCostos = db.session.query(kCentroCostos).all()

        for CentroCosto in CentroCostos:
            if CentroCosto.CodigoPuesto is not None:
                PuestoCentroCosto = db.session.query(tPuesto).filter_by(CodigoPuesto = CentroCosto.CodigoPuesto).first()
                PuestoCentroCosto.idCentroCosto = CentroCosto.idCentroCosto
                Puestos = db.session.query(tPuesto).filter_by(PuestoJefe = CentroCosto.CodigoPuesto, idCentroCosto = None).all()
                for puesto in Puestos:
                    puesto.idCentroCosto = CentroCosto.idCentroCosto

        db.session.commit()

        cont = 1
        Puestos = db.session.query(tPuesto).filter_by(idCentroCosto = None).all()
        while len(Puestos) > 0 and cont < 12:
        
            for puesto in Puestos:
                PuestoCentroCosto = db.session.query(tPuesto).filter_by(CodigoPuesto = puesto.PuestoJefe).first()
                if PuestoCentroCosto is not None:
                    puesto.idCentroCosto = PuestoCentroCosto.idCentroCosto

            db.session.commit()

            Puestos = db.session.query(tPuesto).filter_by(idCentroCosto = None).all()
            cont = cont + 1
            print(cont)

        return({"guardado": True})
    
    else:
        return({"guardado": False})