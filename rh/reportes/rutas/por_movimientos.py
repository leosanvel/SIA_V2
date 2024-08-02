from flask import render_template, request, jsonify, current_app, send_from_directory, url_for
import openpyxl
import openpyxl.workbook
from datetime import datetime, date, time
import os
from dateutil.relativedelta import relativedelta

from .reportes import reportes
from app import db
from rh.gestion_empleados.modelos.empleado import rEmpleadoPuesto, rMovimientoEmpleado, tPuestoHonorarios
from rh.gestion_empleados.modelos.domicilio import rDomicilio
from catalogos.modelos.modelos import kCentroCostos
from general.herramientas.funciones import calcular_quincena

@reportes.route("/rh/reportes/por-movimientos", methods = ["POST", "GET"])
def por_movimientos():

    return render_template("/por_movimientos.html", title = "Por movimientos")

@reportes.route("/rh/reportes/generar_reporte_por_movimiento", methods = ["POST"])
def generar_reporte():
    movimiento = request.form.get("Movimiento")
    wb = openpyxl.Workbook()
    archivo_generado = None
    datos_a_escribir = {}

    dir = os.path.join(current_app.root_path, "rh", "reportes", "archivos", "movimientos")
    if not os.path.exists(dir):
        os.mkdir(dir)
        print("Directorio %s creado" % dir)
    else:
        print("Directorio %s ya existe" % dir)
    
    if movimiento == "1":
        altas = db.session.query(rMovimientoEmpleado).filter(rMovimientoEmpleado.idTipoMovimiento.in_([1, 2])).all()
        for empleado_alta in altas:
            ws = openpyxl.load_workbook(filename="rh/reportes/archivos/PLANTILLA NOMBRAMIENTO ADMINISTRATIVO.xlsx")
            plantilla = ws.active

            empleado = db.session.query(rEmpleadoPuesto).filter_by(idPersona = empleado_alta.idPersonaMod).first()
            domicilio = db.session.query(rDomicilio).filter_by(idPersona = empleado_alta.idPersonaMod, idTipoDomicilio = 1).first()
            if empleado:
                TipoEmpleado = empleado.Empleado.idTipoEmpleado
                hoja = wb.active
                datos_a_escribir["FOLIO"] = empleado_alta.idMovimientoEmpleado
                datos_a_escribir["FECHA"] = datetime.now().strftime("%d-%m-%Y")
                datos_a_escribir["APPATERNO"] = empleado.Empleado.Persona.ApPaterno
                datos_a_escribir["APMATERNO"] = empleado.Empleado.Persona.ApMaterno
                datos_a_escribir["NOMBRE"] = empleado.Empleado.Persona.Nombre
                datos_a_escribir["RFC"] = empleado.Empleado.Persona.RFC if not None else ""
                datos_a_escribir["CURP"] = empleado.Empleado.Persona.CURP
                if empleado.Empleado.Persona.EstadoCivil:
                    datos_a_escribir["ESTADOCIVIL"] = empleado.Empleado.Persona.EstadoCivil.EstadoCivil
                else:
                    datos_a_escribir["ESTADOCIVIL"] = ""
                if empleado.Empleado.Persona.Nacionalidad:
                    datos_a_escribir["NACIONALIDAD"] = empleado.Empleado.Persona.Nacionalidad.Nacionalidad
                else:
                    datos_a_escribir["NACIONALIDAD"] = ""
                datos_a_escribir["SEXO"] = empleado.Empleado.Persona.Sexo
                FechaNacimiento = empleado.Empleado.Persona.FechaNacimiento
                FechaNacimiento = datetime.combine(FechaNacimiento, time())
                FechaActual = datetime.today()
                edad = relativedelta(FechaActual, FechaNacimiento).years
                datos_a_escribir["EDAD"] = edad

                if domicilio is not None:
                    datos_a_escribir["CALLE"] = str(domicilio.Vialidad) + str(domicilio.NumExterior)
                    datos_a_escribir["COLONIA"] = domicilio.idAsentamiento
                    datos_a_escribir["CP"] = str(domicilio.idCP)
                    datos_a_escribir["MUNICIPIO"] = domicilio.Municipio.Municipio
                    datos_a_escribir["ENTIDAD"] = domicilio.Entidad.Entidad
                else:
                    datos_a_escribir["CALLE"] = ""
                    datos_a_escribir["COLONIA"] = ""
                    datos_a_escribir["CP"] = ""
                    datos_a_escribir["MUNICIPIO"] = ""
                    datos_a_escribir["ENTIDAD"] = ""

                
                datos_a_escribir["TELEFONO"] = empleado.Empleado.Persona.TelCasa

                if TipoEmpleado == 1:
                    Puesto = db.session.query(tPuestoHonorarios).filter_by(idPuestoHonorarios = empleado.idPuesto).first()
                    CentroCosto = db.session.query(kCentroCostos).filter_by(idCentroCosto = empleado.idCentroCosto).first()
                    datos_a_escribir["CC"] = CentroCosto.CentroCosto
                    datos_a_escribir["CLAVEPRESUPUESTAL"] = ""
                    datos_a_escribir["NIVEL"] = Puesto.Nivel
                else:
                    datos_a_escribir["CC"] = empleado.Puesto.CentroCostos.CentroCosto
                    datos_a_escribir["CLAVEPRESUPUESTAL"] = empleado.Puesto.CodigoPresupuestal
                    datos_a_escribir["NIVEL"] = empleado.Puesto.NivelSalarial

                FechaInicio = datetime.combine(empleado.FechaInicio, datetime.min.time())
                FechaInicio = FechaInicio.strftime("%d-%m-%Y")
                datos_a_escribir["FECHAINICIO"] = FechaInicio

                for row in plantilla.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            for key, value in datos_a_escribir.items():
                                cell.value = cell.value.replace(f"%{key}%", str(value))

                nombre_archivo = "Nombramiento_" + str(empleado.Empleado.NumeroEmpleado) + ".xlsx"
                ws.save("rh/reportes/archivos/movimientos/Nombramiento_" + str(empleado.Empleado.NumeroEmpleado) + ".xlsx")

        if len(altas) > 0:
            respuesta = True
        else:
            respuesta = False

    if movimiento == "2":
        bajas = db.session.query(rMovimientoEmpleado).filter(rMovimientoEmpleado.idTipoMovimiento == 3).all()
        print(bajas)
        for empleado_baja in bajas:
            empleado = db.session.query(rEmpleadoPuesto).filter_by(idPersona = empleado_baja.idPersonaMod).first()
            ws = openpyxl.load_workbook(filename="rh/reportes/archivos/PLANTILLA FORMATO AVISO DE BAJA.xlsx")
            plantilla = ws.active

            if empleado:
                TipoEmpleado = empleado.Empleado.idTipoEmpleado
                hoja = wb.active

                datos_a_escribir["FOLIO"] = empleado_baja.idMovimientoEmpleado
                datos_a_escribir["APPATERNO"] = empleado.Empleado.Persona.ApPaterno
                datos_a_escribir["APMATERNO"] = empleado.Empleado.Persona.ApMaterno
                datos_a_escribir["NOMBRE"] = empleado.Empleado.Persona.Nombre
                datos_a_escribir["RFC"] = empleado.Empleado.Persona.RFC
                if TipoEmpleado == 1:
                    Puesto = db.session.query(tPuestoHonorarios).filter_by(idPuestoHonorarios = empleado.idPuesto).first()
                    CentroCosto = db.session.query(kCentroCostos).filter_by(idCentroCosto = empleado.idCentroCosto).first()
                    datos_a_escribir["CLAVEPRESUPUESTAL"] = ""
                    datos_a_escribir["NIVEL"] = Puesto.Nivel
                    datos_a_escribir["CC"] = CentroCosto.CentroCosto
                else:
                    datos_a_escribir["CLAVEPRESUPUESTAL"] = empleado.Puesto.CodigoPresupuestal
                    datos_a_escribir["NIVEL"] = empleado.Puesto.NivelSalarial
                    datos_a_escribir["CC"] = empleado.Puesto.idCentroCosto

                datos_a_escribir["CAUSABAJA"] = empleado.idCausaBaja
                datos_a_escribir["NUMEMPLEADO"] = empleado.Empleado.NumeroEmpleado
                datos_a_escribir["CURP"] = empleado.Empleado.Persona.CURP

                FechaBaja = datetime.combine(empleado.FechaEfecto, datetime.min.time())
                FechaBaja = FechaBaja.strftime("%d-%m-%Y")
                datos_a_escribir["FECHABAJA"] = FechaBaja
                

                nombre_archivo = "Baja_" + str(empleado.Empleado.NumeroEmpleado) + ".xlsx"

                for row in plantilla.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            for key, value in datos_a_escribir.items():
                                cell.value = cell.value.replace(f"%{key}%", str(value))

                ws.save("rh/reportes/archivos/movimientos/Baja_" + str(empleado.Empleado.NumeroEmpleado) + ".xlsx")

        if len(bajas) > 0:
            respuesta = True
        else:
            respuesta = False

    if movimiento == "3":
        todos = db.session.query(rMovimientoEmpleado).all()
        cont = 1
        quincena = calcular_quincena()
        hoja = wb.active
        hoja["A1"] = "MOVIMIENTOS DEL PERSONAL DE PLAZA FEDERAL CORRESPONDIENTES A LA QUINCENA"
        hoja["A2"] = "Cons Qnal"
        hoja["B2"] = "No Qna"
        hoja["C2"] = "No Empleado"
        hoja["D2"] = "RFC"
        hoja["E2"] = "CURP"
        hoja["F2"] = "Nombre del Candidato"
        hoja["G2"] = "No Mov"
        hoja["H2"] = "Mov"
        hoja["I2"] = "Plaza"
        hoja["J2"] = "No Plaza"
        hoja["K2"] = "Nivel"
        hoja["L2"] = "CC Origen"
        hoja["M2"] = "Tipo Plaza"
        hoja["N2"] = "Puesto"
        hoja["O2"] = "Efectos a partir de"
        hoja["P2"] = "CC Propuesto"
        hoja["Q2"] = "Centro de Costo"
        hoja["R2"] = "Causa de la Baja"
        hoja["S2"] = "Grupo"
        hoja["T2"] = "Observaciones"
        
        ws = openpyxl.load_workbook(filename="rh/reportes/archivos/PLANTILLA REPORTE DE MOVIMIENTOS.xlsx")
        plantilla = ws.active
        for empleado_comb in todos:
            empleado = db.session.query(rEmpleadoPuesto).filter_by(idPersona = empleado_comb.idPersonaMod).first()
            TipoEmpleado = empleado.Empleado.idTipoEmpleado
            if empleado:
                plantilla["A" + str(6 + cont)] = cont
                plantilla["B" + str(6 + cont)] = str(datetime.now().year) +  str(quincena)
                plantilla["C" + str(6 + cont)] = empleado.Empleado.NumeroEmpleado
                plantilla["D" + str(6 + cont)] = empleado.Empleado.Persona.RFC
                plantilla["E" + str(6 + cont)] = empleado.Empleado.Persona.CURP
                plantilla["F" + str(6 + cont)] = empleado.Empleado.Persona.ApPaterno + " " + empleado.Empleado.Persona.ApMaterno + " " + empleado.Empleado.Persona.Nombre
                plantilla["G" + str(6 + cont)] = empleado_comb.idMovimientoEmpleado
                plantilla["H" + str(6 + cont)] = empleado_comb.idTipoMovimiento
                if TipoEmpleado == 2:
                    plantilla["I" + str(6 + cont)] = ""
                    plantilla["J" + str(6 + cont)] = ""
                    plantilla["K" + str(6 + cont)] = empleado.Puesto.NivelSalarial
                    plantilla["L" + str(6 + cont)] = ""
                    plantilla["M" + str(6 + cont)] = ""
                    plantilla["N" + str(6 + cont)] = empleado.Puesto.Puesto
                    plantilla["O" + str(6 + cont)] = ""
                    plantilla["P" + str(6 + cont)] = ""
                    plantilla["Q" + str(6 + cont)] = empleado.Puesto.CentroCostos.CentroCosto
                else:
                    Puesto = db.session.query(tPuestoHonorarios).filter_by(idPuestoHonorarios = empleado.idPuesto).first()
                    CentroCosto = db.session.query(kCentroCostos).filter_by(idCentroCosto = empleado.idCentroCosto).first()
                    plantilla["I" + str(6 + cont)] = ""
                    plantilla["J" + str(6 + cont)] = ""
                    plantilla["K" + str(6 + cont)] = Puesto.Nivel
                    plantilla["L" + str(6 + cont)] = ""
                    plantilla["M" + str(6 + cont)] = ""
                    plantilla["N" + str(6 + cont)] = Puesto.PuestoHonorarios
                    plantilla["O" + str(6 + cont)] = ""
                    plantilla["P" + str(6 + cont)] = ""
                    plantilla["Q" + str(6 + cont)] = CentroCosto.CentroCosto
                plantilla["R" + str(6 + cont)] = ""
                plantilla["S" + str(6 + cont)] = ""
                plantilla["T" + str(6 + cont)] = empleado.Observaciones
                cont = cont + 1

        ws.save("rh/reportes/archivos/movimientos/Movimientos_" + str(datetime.now().year) +  str(quincena) + ".xlsx")

        nombre_archivo = "Movimientos_" + str(datetime.now().year) +  str(quincena) + ".xlsx"

        if len(todos) > 0:
            respuesta = True
        else:
            respuesta = False

    if respuesta:
        #wb.save(dir + "/" + nombre_archivo)

        archivo_generado = nombre_archivo
    
    else:
        archivo_generado = ""

    return jsonify({"url_descarga": url_for("reportes.descargar_reporte", nombre_archivo=archivo_generado), "respuesta": respuesta})

@reportes.route("/rh/reportes/descargar-reporte/<nombre_archivo>")
def descargar_reporte(nombre_archivo):
    dir = os.path.join(current_app.root_path, "rh", "reportes", "archivos", "movimientos")
    return send_from_directory(directory=dir, path=nombre_archivo, as_attachment=True)