from flask import render_template, request, jsonify, current_app, send_from_directory, url_for
import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime, time
import os

from .reportes import reportes_prestaciones
from app import db
from catalogos.modelos.modelos import kQuincena
from rh.gestion_empleados.modelos.empleado import rEmpleadoPuesto, rEmpleado, tPersona
from prestaciones.modelos.modelos import rEmpleadoConcepto

@reportes_prestaciones.route("/prestaciones/reportes/cifras-control", methods = ['POST', 'GET'])
def cifras_control():
    Quincenas = db.session.query(kQuincena).all()

    return render_template("/cifras_control.html", title = "Cifras de control",
                           Quincenas = Quincenas)

@reportes_prestaciones.route("/prestaciones/reportes/generar-reporte-cifras-control", methods = ["POST"])
def generar_reporte_cifras_control():
    Quincena = request.form.get("Quincenas")
    quincena_reg = db.session.query(kQuincena).filter_by(idQuincena = int(Quincena)).first()
    archivo_generado = None
    anio = datetime.now().year
    quincena = str(anio) + Quincena
    cont = 1

    dir = os.path.join(current_app.root_path, "prestaciones", "reporte", "archivos", "cifras_control")
    if not os.path.exists(dir):
        os.mkdir(dir)
        print("Directorio %s creado" % dir)
    else:
        print("Directorio %s ya existe" % dir)

    wb = openpyxl.Workbook()
    ws = wb.active

    # Asignar titula a la hoja activa
    ws.title = "REPORTE CIFRAS DE CONTROL"

    ws2 = wb.create_sheet(title = "nomina prueba")
    ws3 = wb.create_sheet(title = "NOMINA")
    ws4 = wb.create_sheet(title = "NOMINA PRN")

    encabezado_nomina_prueba = ["RAMO", "PAGADURÍA", "No. ISSSTE", "RFC", "SALARIO BÁSICO", "NOMBRE", "CLAVE DE COBRO", "NOMBRAMIENTO", 'TIPO DE NÓMINA "ORDINARIA"', "SIN USO", "QNA.FECHA INICIO (AAAAMMDD)", "QNA.FECHA FINAL (AAAAMMDD)",
                                "APORTACIONES", "CUOTA SEG.SALUD (42 A Y B)", "CUOTA PREST. (102, 140 Y 199)", "DCTOS. ADICIONALES", "SIN USO", "PRÉSTAMOS PERSONALES ISSSTE", "ADEUDO SERV.MED.", "CVE. DCTO. VIV.", "DCTO. FOVISSSTE",
                                "CVE. DCTO. SEG. DA", "CURP", "SIN USO", "SUMA TOTAL CUOTAS Y DCTOS.", "QNA. (AÑO-QNA)", "FECHA PROC. DEL REPORTE (AAAAMMDD)", "ENTIDAD PAGO", "NSS", "SUELDO SAR", "SIN USO", "TIPO REGISTRO", "",
                                "CONTADOR DECUENTOS CRÉDITOS ISSSTE", "CONTADOR DECUENTOS SEGURO DE DAÑOS", "CONTADOR DECUENTOS CRÉDITOS FOVISSSTE", "CONTADOR TRABAJADORES CON NSS"]
    ws2.append(encabezado_nomina_prueba)

    encabezado_NOMINA = ["", "", "", "", "", "sdo base", "", "ISR", "", "", "", "", "", "PREST PERSONAL", "", "AYUDA SERV", "DESPENSA", "", "", "SEG SALUD", "SEG SALUD PEN", "", "", "SEG VID COLEC", "", "", "", "", "", "FOVISSSTE",
                         "SEG DAÑOS FOV", "SUELDO BASE", "", "SEG AUTO", "", "", "", "", "SEG COL RETIRO", "CAPACITACIÓN", "", "", "", "", "", "FONACOT", "AYUDA TRANSP", "AHORRO SOLID", "RCV", "INV y VIDA", "SEG SOCIAL Y CUL", "", "", "",
                         "", "", "", "", "", "", "", "", "", "", "", "", "", "SUMA QUINQUENIO", "SUELDO COTIZACIÓN CUOTA", "NUMERO TRABAJADORES", "NUM. TRAB. CON SUELDO", "NUM. TRAB. CON NSS", "PRÉSTAMOS ISSSTE", "PRÉSTAMOS FOVISSSTE", "SEGURO DE DAÑOS"]
    ws3.append(encabezado_NOMINA)

    encabezado_NOMINA = ["NRFC", "NCURP", "NNOMBRE", "NNIVEL", "NTIPOPZA", "NSBC_ISSSTE", "NCDT", "NC1", "NC1F", "NC1S", "NC2", "NC21", "NC26", "NC3", "NC32", "NC34", "NC38", "NC4", "NC40", "NC42A", "NC42B", "NC49", "NC5", "NC50", "NC51",
                         "NC55", "NC56L", "NC57", "NC62", "NC64", "NC65", "NC7", "NC71", "NC72", "NC73", "NC74", "NC75", "NC76", "NC77D", "NC77", "NC8", "NC81", "NC82", "NC83", "NC88", "NC94", "NC95", "NC100", "NC102", "NC140", "NC199",
                         "NCCG", "NCDSU", "NCPP", "NCPUV", "NCRET", "NCFT7", "NCFTC", "NCRE7", "NCREC", "NCA1", "NCA2", "NCA3", "NCA4", "NCA5", "NSS", "CVE_PRESUP"]
    ws3.append(encabezado_NOMINA)

    Empleados = db.session.query(rEmpleadoPuesto).join(rEmpleado).join(tPersona).filter(rEmpleado.idTipoEmpleado == 2, rEmpleadoPuesto.idEstatusEP == 1).order_by(tPersona.ApPaterno).all()

    concepto_042A = 0.0
    concepto_042B = 0.0
    concepto_102 = 0.0
    concepto_140 = 0.0
    concepto_199 = 0.0
    sueldos = 0.0

    Conceptos_valores = {}

    num_conceptos = [
    "1", "1F", "1S", "2", "21", "26", "3", "32", "34", "38", "4", "40", "42A", "42B", "49", "5", "50", 
    "51", "55", "56L", "57", "62", "64", "65", "71", "72", "73", "74", "75", "76", "77D", "77", "8", 
    "81", "82", "83", "88", "94", "95", "100", "102", "140", "199", "CG", "DSU", "PP", "PUV", "RET",
    "FT7", "FTC", "RE7", "REC", "A1", "A2", "A3", "A4", "A5"]

    tipo_conceptos = [
        "D", "D", "D", "D", "D", "D", "D", "P", "P", "P", "D", "P", "D", "D", "P", "P", "D", "D", "D", "D",
        "D", "D", "D", "D", "D", "D", "D", "D", "D", "P", "D", "P", "D", "D", "D", "D", "D", "D", "P", "D",
        "D", "D", "D", "P", "P", "P", "P", "P", "D", "D", "P", "P", "P", "P", "P", "P", "P"
    ]

    for empleado in Empleados:
        form_sum_quinquenio = "=SUM(BI" + str(cont + 2) + ":BM" + str(cont + 2) + ")"
        Sueldo = db.session.query(rEmpleadoConcepto).filter_by(idPersona = empleado.idPersona, idTipoConcepto = "P", idConcepto = "7").first()
        sueldos = float(Sueldo.Monto) + sueldos

        for indice in range(0, len(num_conceptos)):
            Concepto = db.session.query(rEmpleadoConcepto).filter_by(idPersona = empleado.idPersona, idTipoConcepto = tipo_conceptos[indice], idConcepto = num_conceptos[indice]).first()
            if Concepto:
                if float(Concepto.Porcentaje) != 0.00:
                    Conceptos_valores[num_conceptos[indice]] = float(Sueldo.Monto*(Concepto.Porcentaje/100))
                else:
                    Conceptos_valores[num_conceptos[indice]] = Concepto.Monto
            else:
                Conceptos_valores[num_conceptos[indice]] = 0.00


        concepto_042A = Conceptos_valores["42A"] + concepto_042A

        concepto_042B = Conceptos_valores["42B"] + concepto_042B

        concepto_102 = Conceptos_valores["102"] + concepto_102

        concepto_140 = Conceptos_valores["140"] + concepto_140

        concepto_199 = Conceptos_valores["199"] + concepto_199

        ws2.append([47, 999030, 0, empleado.Empleado.Persona.RFC, Sueldo.Monto, empleado.Empleado.Persona.ApPaterno + " " + empleado.Empleado.Persona.ApMaterno + " " + empleado.Empleado.Persona.Nombre,
                    empleado.ClavePresupuestaSIA, empleado.Empleado.idTipoEmpleado, 1, "", quincena_reg.FechaInicio, quincena_reg.FechaFin, 111110000, Conceptos_valores["42A"] + Conceptos_valores["42B"],
                    Conceptos_valores["102"] + Conceptos_valores["140"] + Conceptos_valores["199"], 0, 0, "", ""])
        
        ws3.append([empleado.Empleado.Persona.RFC, empleado.Empleado.Persona.CURP, empleado.Empleado.Persona.ApPaterno + " " + empleado.Empleado.Persona.ApMaterno + " " + empleado.Empleado.Persona.Nombre, empleado.idNivel,
                    "P", Sueldo.Monto, 15, Conceptos_valores["1"], Conceptos_valores["1F"], Conceptos_valores["1S"], Conceptos_valores["2"], Conceptos_valores["21"], Conceptos_valores["26"], Conceptos_valores["3"],
                    Conceptos_valores["32"], Conceptos_valores["34"], Conceptos_valores["38"], Conceptos_valores["4"], Conceptos_valores["40"], Conceptos_valores["42A"], Conceptos_valores["42B"], Conceptos_valores["49"],
                    Conceptos_valores["5"], Conceptos_valores["50"], Conceptos_valores["51"], Conceptos_valores["55"], Conceptos_valores["56L"], Conceptos_valores["57"], Conceptos_valores["62"], Conceptos_valores["64"],
                    Conceptos_valores["65"], Sueldo.Monto, Conceptos_valores["71"], Conceptos_valores["72"], Conceptos_valores["73"], Conceptos_valores["74"], Conceptos_valores["75"], Conceptos_valores["76"],
                    Conceptos_valores["77D"], Conceptos_valores["77"], Conceptos_valores["8"], Conceptos_valores["81"], Conceptos_valores["82"], Conceptos_valores["83"], Conceptos_valores["88"], Conceptos_valores["94"],
                    Conceptos_valores["95"], Conceptos_valores["100"], Conceptos_valores["102"], Conceptos_valores["140"], Conceptos_valores["199"], Conceptos_valores["CG"], Conceptos_valores["DSU"], Conceptos_valores["PP"],
                    Conceptos_valores["PUV"], Conceptos_valores["RET"], Conceptos_valores["FT7"], Conceptos_valores["FTC"], Conceptos_valores["RE7"], Conceptos_valores["REC"], Conceptos_valores["A1"], Conceptos_valores["A2"],
                    Conceptos_valores["A3"], Conceptos_valores["A4"], Conceptos_valores["A5"], "", empleado.ClavePresupuestaSIA, form_sum_quinquenio])
        
        cont = cont + 1

    # Asiganar los datos
    data = [
        ["REPORTE DE CIFRAS DE CONTROL DE LA INFORMACIÓN DE NÓMINA"],
        [],
        ["Ramo 99903", "", "Pagaduria", "", "00047", "", "", "", "", "QNA.", quincena],
        ["Nombre", "", "Instituto Nacional de la Economia Social (INAES)", "", ""],
        [],
        ["CONCEPTO", "Art. Ley", "%", "No. REG", "NÓMINA ORDINARIA", "No. REG", "PAGOS CANCELADOS", "No. REG", "NÓMINA EXTRAORDINARIA", "TOTAL REG.", "TOTAL"],
        [],
        [],
        ["IMPORTE TOTAL DE SUELDOS", "", "", len(Empleados), sueldos, "", "", "", "", "=SUM(D9+F9+H9)", "=SUM(E9+G9+I9)"],
        [],
        ["CUOTAS DE LOS TRABAJADORES"],
        [],
        ["SEGURO DE SALUD DE LOS TRABAJADORES EN ACTIVO Y FAMILIARES DERECHOHABIENTES", "042A", "2.75%", len(Empleados), concepto_042A, "", "", "", "", "=SUM(D13+F13+H13)", "=SUM(E13+G13+I13)"],
        ["SEGURO DE SALUD DE LOS PENSIONISTAS Y FAMILIARES DERECHOHABIENTES", "042B", "6.25%", len(Empleados), concepto_042B, "", "", "", "", "=D14+F14+H14", "=SUM(E14+G14+I14)"],
        ["RETIRO POR EDAD AVANZADA Y VEJEZ", "102", "2%", len(Empleados), concepto_102, "", "", "", "", "=D15+F15+H15", "=SUM(E15+G15+I15)"],
        ["SEGURO DE INVALIDEZ Y VIDA", "140", "2%", len(Empleados), concepto_140, "", "", "", "", "=D16+F16+H16", "=SUM(E16+G16+I16)"],
        ["SERVICIOS SOCIALES Y CULTURALES", "199", "0.50%", len(Empleados), concepto_199, "", "", "", "", "=D16+F16+H16", "=SUM(E17+G17+I17)"],
        [],
        ["", "", "", "=D9", "=SUM(E13:E17)", "", "", "", "", "=J9", "=SUM(K13:K17)"],
        [],
        ["DESCUENTOS A TRABAJADORES"],
        [],
        ["PRESTAMOS A CORTO PLAZO", "", "", 95, "$ 87,062.35", "", "", "", "", 95, "$ 87,062.35"],
        ["ADEUDOS S. MEDICOS", "", "", "", "", "", "", "", "", "", ""],
        ["PRESTAMO HIPOTECARIO", "", "", 164, "$ 1,394.00", "", "", "", "", 164, "$ 1,394.00"],
        ["SEGURO HIPOTECARIO", "", "", 164, "", "", "", "", "", "", ""],
        ["PRESTAMOS ADICIONALES", "", "", "", "", "", "", "", "", "", ""],
        ["SEGURO HIPOT. AVALADO", "", "", "", "", "", "", "", "", "", ""],
        ["FOVISSSTE CONSTANTE", "", "", "", "", "", "", "", "", "", ""],
        ["FOVISSSTE CRECIENTE", "", "", "", "", "", "", "", "", "", ""],
        ["FOVISSSTE SAL. MIN.", "167", "", 167, "$ 216,320.17", "", "", "", "", 167, "$ 216,320.17"],
        ["IMPORTE SAR", "", "", "", "", "", "", "", "", "", ""],
        [],
        ["TOTAL (CUOTAS + DESCUENTOS)", "", "", 352, "$ 477,228.32", "", "", "", "", 352, "$ 477,228.32"],
        [],
        ["NUMERO DE TRABAJADORES", "", "", "", "", "", "", "", "", "", ""],
        ["NUMERO DE REGISTROS CON CURP", "", "", 352, "", "", "", "", "", "", ""],
        ["NUMERO DE REGISTROS CON NSS", "", "", 346, "", "", "", "", "", "", ""],
    ]

    # agregar datos a hoja activa
    for row in data:
        ws.append(row)

    # Merge cells for the title and other header rows
    ws.merge_cells('A1:K1')
    #ws.merge_cells('A2:A3')
    #ws.merge_cells('B2:E2')
    #ws.merge_cells('F2:G2')
    #ws.merge_cells('H2:I2')
    #ws.merge_cells('J2:K2')

    # Apply styles for header rows
    header_font = Font(bold=True)
    alignment = Alignment(horizontal='center', vertical='center')

    for cell in ws["A1:K1"][0]:
        cell.font = header_font
        cell.alignment = alignment

    for cell in ws["A2:K2"][0]:
        cell.font = header_font
        cell.alignment = alignment

    for cell in ws["A4:K4"][0]:
        cell.font = header_font
        cell.alignment = alignment

    # Save the workbook to a file
    file_path = "prestaciones/reporte/archivos/cifras_control/PROCESO DE HOJA DE CIFRAS DE CONTROL.xlsx"
    wb.save(file_path)

    print(f'Excel file created at {file_path}')

    return jsonify({"guardado": True})
        