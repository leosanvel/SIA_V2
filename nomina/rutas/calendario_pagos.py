from flask import render_template, request, jsonify, url_for
from datetime import date, datetime
import openpyxl
from openpyxl.styles import NamedStyle, Border, Side, Alignment

from app import db
from .rutas import nomina
from nomina.modelos.modelos import kMeses, kQuincenaCalendario, kActividadCalendario, tFechasCalendario

@nomina.route("/nomina/calendario-pagos", methods = ['GET', 'POST'])
def calendario_pagos():
    Meses = db.session.query(kMeses).order_by(kMeses.idMes).all()
    QuincenasCalendario = db.session.query(kQuincenaCalendario).filter(kQuincenaCalendario.idQuincenaCalendario.in_([1, 2])).order_by(kQuincenaCalendario.idQuincenaCalendario).all()
    Actividades = db.session.query(kActividadCalendario).order_by(kActividadCalendario.idActividadCalendario).all()

    Anio = date.today().year

    return render_template("/calendario_pagos.html", title = "Calendario de pagos",
                           Meses = Meses,
                           QuincenasCalendario = QuincenasCalendario,
                           Actividades = Actividades,
                           Anio = Anio)

@nomina.route("/nomina/obtener-select-concepto", methods = ['POST'])
def obtener_select_concepto():
    idMes = request.form.get('idMes')
    if(idMes != '12'):
        quincenas = db.session.query(kQuincenaCalendario).filter(kQuincenaCalendario.idQuincenaCalendario.in_([1, 2])).order_by(kQuincenaCalendario.idQuincenaCalendario).all()
    else:
        quincenas = db.session.query(kQuincenaCalendario).filter(kQuincenaCalendario.idQuincenaCalendario.in_([3, 4])).order_by(kQuincenaCalendario.idQuincenaCalendario).all()

    ret = '<option value="0">-- Seleccione --</option>'

    for registro in quincenas:
        ret += '<option value="{}">{}</option>'.format(registro.idQuincenaCalendario, registro.QuincenaCalendario)

    return ret

@nomina.route("/nomina/agregar-fecha-calendario", methods = ['POST'])
def agregar_fecha_calendario():
    mapeo_nombres = { #NombreEnFormulario : nombreEnBase
        'Mes': 'idMes',
        'Concepto': 'idQuincenaCalendario',
        'Actividad': 'idActividadCalendario',
        'FechaInicio': 'FechaInicio',
        'FechaFin': 'FechaFin',
        'Anio': 'Periodo'
    }

    fechas_a_agregar = {mapeo_nombres[key]: request.form.get(key) for key in mapeo_nombres.keys()}
    fechas_a_agregar["FechaInicio"] = datetime.strptime(fechas_a_agregar["FechaInicio"], '%d/%m/%Y')
    fechas_a_agregar["FechaFin"] = datetime.strptime(fechas_a_agregar["FechaFin"], '%d/%m/%Y')
    nuevo_fecha_calendario = None
    fechas_calendario_existente = db.session.query(tFechasCalendario).filter_by(idMes = fechas_a_agregar["idMes"], idQuincenaCalendario = fechas_a_agregar["idQuincenaCalendario"], idActividadCalendario = fechas_a_agregar["idActividadCalendario"], Periodo = fechas_a_agregar["Periodo"]).first()
    if fechas_calendario_existente is None:
        nuevo_fecha_calendario = tFechasCalendario(**fechas_a_agregar)
        db.session.add(nuevo_fecha_calendario)
        db.session.commit()
        return jsonify({"guardado": True})
    else:
        fechas_calendario_existente.update(**fechas_a_agregar)
        db.session.commit()
        return jsonify({"actualizado": True})

@nomina.route("/nomina/obtener-fechas-calendario", methods = ["POST"])
def buscar_fechas_calendario():
    anio = date.today().year
    Fechas = db.session.query(tFechasCalendario).filter_by(Periodo = anio).order_by(tFechasCalendario.idMes, tFechasCalendario.idActividadCalendario).all()
    Meses = db.session.query(kMeses).order_by(kMeses.idMes).all()
    QuincenasCalendario = db.session.query(kQuincenaCalendario).filter_by(Activo = 1).order_by(kQuincenaCalendario.idQuincenaCalendario).all()
    Actividades = db.session.query(kActividadCalendario).order_by(kActividadCalendario.idActividadCalendario).all()

    Lista_Meses = []
    Lista_QuincenasCalendario = []
    Lista_Actividades = []
    Lista_Fechas = []

    for mes in Meses:
        mes_dict = mes.__dict__
        mes_dict.pop("_sa_instance_state", None)
        Lista_Meses.append(mes_dict)

    for Quincena in QuincenasCalendario:
        Quincena_dict = Quincena.__dict__
        Quincena_dict.pop("_sa_instance_state", None)
        Lista_QuincenasCalendario.append(Quincena_dict)

    for Actividad in Actividades:
        Actividad_dict = Actividad.__dict__
        Actividad_dict.pop("_sa_instance_state", None)
        Lista_Actividades.append(Actividad_dict)

    for fecha in Fechas:
        fecha_dict = fecha.__dict__
        fecha_dict.pop("_sa_instance_state", None)
        Lista_Fechas.append(fecha_dict)

    return jsonify({'Meses': Lista_Meses,
                    'QuincenasCalendario': Lista_QuincenasCalendario,
                    'Actividades': Lista_Actividades,
                    'Fechas': Lista_Fechas})

@nomina.route("/nomina/generar-calendario-pagos", methods = ["POST"])
def generar_calendario_pagos():
    anio = date.today().year
    Fechas = db.session.query(tFechasCalendario).filter_by(Periodo = anio).all()

    wb = openpyxl.load_workbook(filename="nomina/docs_plantillas/Plantilla_Calendario_de_Pagos.xlsx", rich_text=True)

    #wb.active

    datos_a_insertar = {
        'PERIODO': str(anio)
    }

    hojas = ['1er_Trim', '2do_Trim', '3er_Trim', '4o_Trim']

    estilo_texto_centrado = NamedStyle(name="texto_centrado")
    estilo_texto_centrado.alignment = Alignment(horizontal='center', vertical='center')
    estilo_borde_izq = Border(right=Side(style='thin'))

    for row in wb[hojas[1]].iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                for key, value in datos_a_insertar.items():
                    cell.value = cell.value.replace(f"%{key}%", str(value))

    for fecha in Fechas:
        if fecha.idMes == 1 or fecha.idMes == 2 or fecha.idMes == 3:
            fila = calcular_fila(fecha.idMes, fecha.idQuincenaCalendario)
            columna = calcular_columna(fecha.idActividadCalendario)
            hoja = 0

        elif fecha.idMes == 4 or fecha.idMes == 5 or fecha.idMes == 6:
            fila = calcular_fila(fecha.idMes, fecha.idQuincenaCalendario)
            columna = calcular_columna(fecha.idActividadCalendario)
            hoja = 1

        elif fecha.idMes == 7 or fecha.idMes == 8 or fecha.idMes == 9:
            fila = calcular_fila(fecha.idMes, fecha.idQuincenaCalendario)
            columna = calcular_columna(fecha.idActividadCalendario)
            hoja = 2

        elif fecha.idMes == 4 or fecha.idMes == 5 or fecha.idMes == 6:
            fila = calcular_fila(fecha.idMes, fecha.idQuincenaCalendario)
            columna = calcular_columna(fecha.idActividadCalendario)
            hoja = 3

        else:
            print("Error en el mes.")

        if(fecha.FechaInicio == fecha.FechaFin):
            rango = columna[0] + fila + ':' + columna[2] + fila
            wb[hojas[hoja]].merge_cells(rango)
            wb[hojas[hoja]][columna[0] + fila] = fecha.FechaInicio.day
            wb[hojas[hoja]][columna[0] + fila].style = estilo_texto_centrado

        else:
            wb[hojas[hoja]][columna[0] + fila] = fecha.FechaInicio.day
            wb[hojas[hoja]][columna[0] + fila].style = estilo_texto_centrado
            wb[hojas[hoja]][columna[1] + fila] = '-'
            wb[hojas[hoja]][columna[1] + fila].style = estilo_texto_centrado
            wb[hojas[hoja]][columna[2] + fila] = fecha.FechaFin.day
            wb[hojas[hoja]][columna[2] + fila].style = estilo_texto_centrado
            wb[hojas[hoja]][columna[2] + fila].border = estilo_borde_izq


    wb.save("nomina/documentos/CALENDARIO DE PAGOS " + str(anio) + ".xlsx")

    return({"guardado": True})

def calcular_fila(Mes, Quincena):
    if Mes % 3 == 1 and Quincena == 1:
        fila = '10'
            
    elif Mes % 3 == 1 and Quincena == 2:
        fila = '13'

    elif Mes % 3 == 2 and Quincena == 1:
        fila = '17'

    elif Mes % 3 == 2 and Quincena == 2:
        fila = '20'

    elif Mes % 3 == 0 and Quincena == 1:
        fila = '24'
            
    elif Mes % 3 == 0 and Quincena == 2:
        fila = '27'

    return fila

def calcular_columna(Actividad):
    if Actividad == 1:
        columna = ['D', 'E', 'F']

    elif Actividad == 2:
        columna = ['G', 'H', 'I']

    elif Actividad == 3:
        columna = ['J', 'K', 'L']

    elif Actividad == 4:
        columna = ['M', 'N', 'O']
    
    elif Actividad == 5:
        columna = ['P', 'Q', 'R']

    elif Actividad == 6:
        columna = ['S', 'T', 'U']

    elif Actividad == 7:
        columna = ['V', 'W', 'X']

    return columna