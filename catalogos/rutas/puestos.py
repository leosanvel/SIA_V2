from .rutas import catalogos
from flask import render_template, request, jsonify
from flask_login import current_user
from sqlalchemy.orm.exc import NoResultFound
from sqlalchemy import inspect, or_
import openpyxl
import collections
# import pandas as pd

from app import db
from rh.gestion_empleados.modelos.empleado import tPuesto
from catalogos.modelos.modelos import kRamo, kUA, kZonaEconomica, kTipoPlazaPuesto, kCaracterOcupacional, kTipoFuncion, kGrupo, kGrado, kNivel, kEstatusPuesto, kVigencia, kCentroTrabajo, kCentroCostos

@catalogos.route('/catalogos/puestos')
def catalogos_puestos():
    Ramo = db.session.query(kRamo).filter_by(Activo = 1).order_by(kRamo.idRamo).all()
    UA = db.session.query(kUA).filter_by(Activo = 1).order_by(kUA.idUA).all()
    ZonaEconomica = db.session.query(kZonaEconomica).filter_by(Activo = 1).order_by(kZonaEconomica.idZonaEconomica).all()
    TipoPlazaPuesto = db.session.query(kTipoPlazaPuesto).filter_by(Activo = 1).order_by(kTipoPlazaPuesto.idTipoPlazaPuesto).all()
    CaracterOcupacional = db.session.query(kCaracterOcupacional).filter_by(Activo = 1).order_by(kCaracterOcupacional.idCaracterOcupacional).all()
    TipoFuncion = db.session.query(kTipoFuncion).filter_by(Activo = 1).order_by(kTipoFuncion.idTipoFuncion).all()
    Grupo = db.session.query(kGrupo).filter_by(Activo = 1).order_by(kGrupo.idGrupo).all()
    Grado = db.session.query(kGrado).filter_by(Activo = 1).order_by(kGrado.idGrado).all()
    Nivel = db.session.query(kNivel).filter_by(Activo = 1).order_by(kNivel.idNivel).all()
    EstatusPuesto = db.session.query(kEstatusPuesto).filter_by(Activo = 1).order_by(kEstatusPuesto.idEstatusPuesto).all()
    Vigencia = db.session.query(kVigencia).filter_by(Activo = 1).order_by(kVigencia.idVigencia).all()
    CentroTrabajo = db.session.query(kCentroTrabajo).filter_by(Activo = 1).order_by(kCentroTrabajo.idCentroTrabajo).all()
    CentroCostos = db.session.query(kCentroCostos).order_by(kCentroCostos.idCentroCosto).all()

    return render_template('/puestos.html', title = 'Puestos',
                           Ramo = Ramo,
                           UA = UA,
                           ZonaEconomica = ZonaEconomica,
                           TipoPlazaPuesto = TipoPlazaPuesto,
                           CaracterOcupacional = CaracterOcupacional,
                           TipoFuncion = TipoFuncion,
                           Grupo = Grupo,
                           Grado = Grado,
                           Nivel = Nivel,
                           EstatusPuesto = EstatusPuesto,
                           Vigencia = Vigencia,
                           CentroTrabajo = CentroTrabajo,
                           CentroCostos = CentroCostos)

@catalogos.route("/catalogos/crear-puesto", methods = ["POST"])
def crear_puesto():

    mapeo_nombres = { #NombreEnFormulario : nombreEnBase
        "Ramo": "idRamo",
        "UA": "idUA",
        "ConsecutivoPuesto": "ConsecutivoPuesto",
        "CodigoPuesto": "CodigoPuesto",
        "Puesto": "Puesto",
        "ZonaEconomica": "idZonaEconomica",
        "ReferenciaTabular": "ReferenciaTabular",
        "ConsPuesto": "ConsPuesto",
        "TipoPlazaPuesto": "idTipoPlazaPuesto",
        "CaracterOcupacional": "idCaracterOcupacional",
        "TipoFuncion": "idTipoFuncion",
        "NivelSalarial": "NivelSalarial",
        "Tabulador": "Tabulador",
        "CodigoPresupuestal": "CodigoPresupuestal",
        "OrdinalCP": "OrdinalCP",
        "Grupo": "idGrupo",
        "Grado": "idGrado",
        "Nivel": "idNivel",
        "EstatusPuesto": "idEstatusPuesto",
        "Vigencia": "idVigencia",
        "FechaInicio": "FechaInicio",
        "FechaFin": "FechaFin",
        "CentroTrabajo": "idCentroTrabajo",
        "FolioSival": "FolioSival",
        "RegimenLaboral": "RegimenLaboral",
        "RemuneracionTotal": "RemuneracionTotal",
        "TitularAU": "TitularAU",
        "DeclaracionPatrimonial": "DeclaracionPatrimonial",
        "PlazasSubordinadas": "PlazasSubordinadas",
        "PuestoJefe": "PuestoJefe",
        "PresupuestalJefe": "PresupuestalJefe",
        "CentroCosto": "idCentroCosto"
    }

    puesto_data = {mapeo_nombres[key]: request.form.get(key) for key in mapeo_nombres.keys()}
    nuevo_puesto = None
    print(puesto_data)

    try:
        puesto_modificar = db.session.query(tPuesto).filter_by(ConsecutivoPuesto = puesto_data["ConsecutivoPuesto"]).one()
        puesto_modificar.update(**puesto_data)
        print("MODIFICAR")

    except NoResultFound:
        nuevo_puesto = tPuesto(**puesto_data)
        db.session.add(nuevo_puesto)
        print("AÑADIR NUEVO")

    #db.session.commit()

    return jsonify(puesto_data)

@catalogos.route("/catalogos/actualizar-busqueda-puestos", methods = ["GET"])
def actualizar_busqueda_puestos():
    texto_busqueda = request.args.get("texto_busqueda", "")
    resultados_json = []
    resultados = tPuesto.query.filter(or_(tPuesto.Puesto.ilike(f'%{texto_busqueda}%'), tPuesto.ConsPuesto.ilike(f'%{texto_busqueda}%'))).all()

    for resultado in resultados:
        if resultado is not None:
            resultado_dict = resultado.__dict__
            resultado_dict.pop("_sa_instance_state", None)  # Eliminar atributo de SQLAlchemy
            resultados_json.append(resultado_dict)

    return jsonify(resultados_json)

@catalogos.route("/catalogos/buscar-puesto", methods = ["POST"])
def buscar_puesto():
    respuesta = {}
    puesto = request.form.get("PuestoExistente")
    if puesto:
        puestos = db.session.query(tPuesto).filter(tPuesto.Puesto.contains(puesto)).all()

    else:
        puestos = []

    lista_puestos = []
    for pues in puestos:
        if pues is not None:
            puesto_dict = pues.__dict__
            puesto_dict.pop("_sa_instance_state", None)  # Eliminar atributo de SQLAlchemy
            lista_puestos.append(puesto_dict)
    
    if not lista_puestos:
        respuesta["NoEncontrado"] = True

    respuesta["ListaPuestos"] = lista_puestos

    return jsonify(respuesta)

@catalogos.route("/catalogos/cargar-archivo-puestos", methods = ["POST"])
def cargar_archivo_puestos():
    archivo = request.files.get('archivo')
    #print(archivo)

    columnas = inspect(tPuesto).all_orm_descriptors.keys()
    del columnas[32:44]

    valor_columnas =[1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 32, 41, 40]
    mapeo_nombres = dict(zip(columnas, valor_columnas))

    if(archivo):
        wb = openpyxl.load_workbook(archivo)
        print(wb)
        ws = wb.active

        datos_lista = []
        datos_dict = {}
        nuevo_puesto = None

        #idEstatusPuesto, idVigencia, idCentroTrabajo

        for row_value in range(12, ws.max_row + 1):
            for key in mapeo_nombres:
                datos_dict.update({key: ws.cell(row = row_value, column = mapeo_nombres[key]).value})
                if key == "idEstatusPuesto":
                    valor = db.session.query(kEstatusPuesto).filter_by(EstatusPuesto = datos_dict["idEstatusPuesto"]).one()
                    datos_dict["idEstatusPuesto"] = valor.idEstatusPuesto
                
                if key == "idVigencia":
                    valor = db.session.query(kVigencia).filter_by(Vigencia = datos_dict["idVigencia"]).one()
                    datos_dict["idVigencia"] = valor.idVigencia

                if key == "idCentroTrabajo":
                    valor = db.session.query(kCentroTrabajo).filter_by(CentroTrabajo = datos_dict["idCentroTrabajo"]).one()
                    datos_dict["idCentroTrabajo"] = valor.idCentroTrabajo

            datos_lista.append(datos_dict.copy())

            try:
                puesto_existente = db.session.query(tPuesto).filter_by(ConsecutivoPuesto = datos_dict["ConsecutivoPuesto"]).one()
                puesto_existente.update(**datos_dict)

            except NoResultFound:
                nuevo_puesto = tPuesto(**datos_dict)
                db.session.add(nuevo_puesto)

        db.session.commit()


        return({"guardado": True})
    
    else:
        return({"guardado": False})